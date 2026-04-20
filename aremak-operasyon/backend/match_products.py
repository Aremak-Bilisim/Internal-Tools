import openpyxl, re
from difflib import SequenceMatcher
from openpyxl.styles import Font, PatternFill, Alignment

# Load TeamGram
wb_tg = openpyxl.load_workbook('C:/Users/balam/Downloads/Tum urunler_2026-04-18-10-55.xlsx')
ws_tg = wb_tg.active
raw_headers = [str(c.value or '') for c in next(ws_tg.iter_rows(min_row=1, max_row=1))]
print("TG headers:", raw_headers)

tg_products = []
for row in ws_tg.iter_rows(min_row=2, values_only=True):
    d = dict(zip(raw_headers, row))
    marka = str(d.get('Marka') or '').strip()
    model = ''
    isim = ''
    kod = ''
    for k, v in d.items():
        kl = k.lower()
        if 'model' in kl:
            model = str(v or '').strip()
        if 'sim' in kl and 'marka' not in kl and 'birim' not in kl and 'kri' not in kl:
            isim = str(v or '').strip()
        if 'kodu' in kl:
            kod = str(v or '').strip()
    display = model if model else isim
    tg_products.append({'id': d.get('Id'), 'marka': marka, 'model': model, 'isim': isim, 'kod': kod, 'display': display})

# Load Parasut (from official Excel export — columns: name, code, barkod, ...)
wb_ps = openpyxl.load_workbook('C:/Users/balam/Downloads/hizmet-ve-urunler (5).xlsx')
ws_ps = wb_ps.active
ps_products = []
for row in ws_ps.iter_rows(min_row=2, values_only=True):
    name = str(row[0] or '').strip()
    code = str(row[1] or '').strip()
    unit = str(row[10] or '').strip()
    if name:
        ps_products.append({'id': '', 'name': name, 'code': code, 'unit': unit})

print(f"TeamGram: {len(tg_products)}, Parasut: {len(ps_products)}")

def normalize(s):
    return re.sub(r'[\s\-_]+', '', (s or '').lower())

ps_by_norm_name = {normalize(p['name']): p for p in ps_products}
ps_by_code = {normalize(p['code']): p for p in ps_products if p['code']}

def extract_arducam_code(model_str):
    """Extract Bxxxx or EKxxx code from Arducam model string like 'B0165 - Arducam ...' or 'EK039 - ...'"""
    m = re.match(r'^((?:B|EK)\d{3,4})', model_str.strip(), re.IGNORECASE)
    return m.group(1).upper() if m else None

def strip_stars(s):
    return re.sub(r'\*+', '', s or '').strip()

def find_match(tg):
    marka = tg['marka']

    # --- TIS: * temizle, sadece tam eşleşme (boşluk farkı normalize edilir) ---
    if marka == 'TIS':
        model_clean = strip_stars(tg['model'])
        isim_clean  = strip_stars(tg['isim'])
        model_norm  = normalize(model_clean)
        isim_norm   = normalize(isim_clean)
        if model_norm and model_norm in ps_by_norm_name:
            return ps_by_norm_name[model_norm], 'Tam (TIS model)'
        if isim_norm and isim_norm in ps_by_norm_name:
            return ps_by_norm_name[isim_norm], 'Tam (TIS isim)'
        return None, 'Eslesme yok'

    model_norm = normalize(tg['model'])
    isim_norm = normalize(tg['isim'])
    kod_norm = normalize(tg['kod'])

    # --- Hikrobot: sadece birebir model eşleşmesi ---
    if marka == 'Hikrobot':
        if model_norm and model_norm in ps_by_norm_name:
            return ps_by_norm_name[model_norm], 'Tam (Hikrobot model)'
        return None, 'Eslesme yok'

    # --- Arducam: B0xxx kodu üzerinden eşleştir ---
    if marka == 'Arducam':
        arc_code = extract_arducam_code(tg['model'])
        if arc_code:
            # Parasut adında "- B0165 -" geçiyorsa
            for ps in ps_products:
                if arc_code in ps['name'].upper():
                    return ps, f'Arducam kod ({arc_code})'
            # Parasut stok kodunda geçiyorsa
            for ps in ps_products:
                if arc_code in ps['code'].upper():
                    return ps, f'Arducam stok kodu ({arc_code})'
        # TG ürün kodu == Parasut kodu
        if kod_norm and kod_norm in ps_by_code:
            return ps_by_code[kod_norm], 'Kod eslesmesi'
        return None, 'Eslesme yok'

    # --- Diger markalar ---
    # 1. Exact model == Parasut name
    if model_norm and model_norm in ps_by_norm_name:
        return ps_by_norm_name[model_norm], 'Tam (model=ad)'
    # 2. Exact isim == Parasut name
    if isim_norm and isim_norm in ps_by_norm_name:
        return ps_by_norm_name[isim_norm], 'Tam (isim=ad)'
    # 3. TG code == Parasut code
    if kod_norm and kod_norm in ps_by_code:
        return ps_by_code[kod_norm], 'Kod eslesmesi'
    # 4. Parasut name contains TG model (exact substring)
    if model_norm and len(model_norm) > 4:
        for ps in ps_products:
            if model_norm in normalize(ps['name']):
                return ps, 'Kismi (model icinde)'
    # 5. Fuzzy (non-camera products only)
    if tg['model'] and marka not in ('The Imaging Source', 'Hikrobot', 'HIKROBOT'):
        best_score, best_match = 0, None
        for ps in ps_products:
            score = SequenceMatcher(None, normalize(tg['model']), normalize(ps['name'])).ratio()
            if score > best_score:
                best_score, best_match = score, ps
        if best_score > 0.85:
            return best_match, f'Benzer (%{int(best_score*100)})'
    return None, 'Eslesme yok'

# Load existing approvals by TG ID to preserve manual edits
out_path = 'C:/Users/balam/Desktop/urun_eslestirme.xlsx'
existing_approvals = {}  # {tg_id: {'onay': ..., 'not': ...}}
import os
if os.path.exists(out_path):
    try:
        wb_existing = openpyxl.load_workbook(out_path)
        ws_existing = wb_existing.active
        ex_headers = [str(c.value or '') for c in ws_existing[1]]
        idx_id   = ex_headers.index('TG ID')   if 'TG ID'               in ex_headers else None
        idx_onay = ex_headers.index('Onay (Evet/Hayir)') if 'Onay (Evet/Hayir)' in ex_headers else None
        idx_not  = ex_headers.index('Not / Duzeltme')    if 'Not / Duzeltme'    in ex_headers else None
        if idx_id is not None:
            for ex_row in ws_existing.iter_rows(min_row=2, values_only=True):
                tg_id = ex_row[idx_id]
                if tg_id is None:
                    continue
                existing_approvals[tg_id] = {
                    'onay': str(ex_row[idx_onay] or '') if idx_onay is not None else '',
                    'not':  str(ex_row[idx_not]  or '') if idx_not  is not None else '',
                }
        print(f"Mevcut onaylar yuklendi: {len(existing_approvals)} kayit")
    except Exception as e:
        print(f"Mevcut dosya okunamadi, onaylar korunamayacak: {e}")

# Build output
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'Eslestirme'

col_headers = [
    'TG ID', 'TG Marka', 'TG Model', 'TG Urun Kodu',
    'Parasut ID', 'Parasut Urun Adi', 'Parasut Stok Kodu',
    'Eslesme Turu', 'Onay (Evet/Hayir)', 'Not / Duzeltme'
]
ws_out.append(col_headers)

header_fill = PatternFill(start_color='1F3A6E', end_color='1F3A6E', fill_type='solid')
for cell in ws_out[1]:
    cell.font = Font(color='FFFFFF', bold=True)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

fill_exact   = PatternFill(start_color='D6FFD6', end_color='D6FFD6', fill_type='solid')
fill_partial = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
fill_none    = PatternFill(start_color='FFD6D6', end_color='FFD6D6', fill_type='solid')

stats = {'tam': 0, 'kismi': 0, 'yok': 0}
for tg in tg_products:
    if not tg['model'] and not tg['isim']:
        continue
    ps, mtype = find_match(tg)
    saved = existing_approvals.get(tg['id'], {})
    onay = saved.get('onay') or ('Evet' if ps else '')
    not_val = saved.get('not') or ''
    display_model = strip_stars(tg['model'] or tg['isim'])
    ws_out.append([
        tg['id'], tg['marka'], display_model, tg['kod'],
        ps['id'] if ps else '', ps['name'] if ps else '', ps['code'] if ps else '',
        mtype, onay, not_val
    ])
    r = ws_out.max_row
    if 'Tam' in mtype or 'Kod' in mtype:
        fill = fill_exact; stats['tam'] += 1
    elif ps:
        fill = fill_partial; stats['kismi'] += 1
    else:
        fill = fill_none; stats['yok'] += 1
    for cell in ws_out[r]:
        cell.fill = fill

widths = [12, 15, 30, 35, 14, 45, 30, 18, 16, 25]
for i, w in enumerate(widths, 1):
    ws_out.column_dimensions[chr(64+i)].width = w

wb_out.save(out_path)
print(f"Kaydedildi: {out_path}")
print(f"Tam: {stats['tam']}, Kismi: {stats['kismi']}, Yok: {stats['yok']}")
