"""
Hikrobot Commercial Invoice (CI) Excel parser.
'CI' sayfasındaki "Model name" + "Quantities (PCS)" kolonlarını çıkarır.
Aynı model birden çok satırda varsa quantity'leri toplar.
"""
import io
from typing import Optional
import openpyxl


def _norm_header(s) -> str:
    if s is None:
        return ""
    return str(s).strip().lower().replace("\n", " ")


def parse_ci_quantities(file_bytes: bytes) -> dict:
    """
    Excel'in 'CI' sayfasından model adı → toplam adet dict'i çıkarır.
    Örnek dönüş:
    {
      "items": [{"model_name": "MV-CU120-10GM", "quantity": 15}, ...],
      "total": 80,
    }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # 'CI' sayfasını bul (varsa, yoksa ilk sayfa)
    ws = None
    for name in wb.sheetnames:
        if name.strip().upper() == "CI":
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    # Header satırını bul: "Model name" ve "Quantities" geçen satır
    header_row = None
    col_model = None
    col_qty = None
    for r in range(1, min(ws.max_row + 1, 50)):
        for c in range(1, ws.max_column + 1):
            h = _norm_header(ws.cell(r, c).value)
            if h == "model name":
                col_model = c
                header_row = r
            elif "quantit" in h:  # "Quantities (PCS)" veya benzeri
                col_qty = c
        if header_row and col_qty:
            break

    if not header_row or not col_model or not col_qty:
        raise ValueError("Excel'de 'Model name' veya 'Quantities' kolonu bulunamadı")

    # Header'dan sonraki satırları oku
    totals: dict[str, float] = {}
    grand_total = 0.0
    for r in range(header_row + 1, ws.max_row + 1):
        model = ws.cell(r, col_model).value
        qty = ws.cell(r, col_qty).value
        if not model:
            continue
        model_str = str(model).strip()
        if not model_str:
            continue
        # TOTAL satırını atla
        if model_str.upper().startswith("TOTAL"):
            continue
        try:
            qty_f = float(qty) if qty is not None else 0
        except (TypeError, ValueError):
            continue
        if qty_f <= 0:
            continue
        totals[model_str] = totals.get(model_str, 0) + qty_f
        grand_total += qty_f

    items = [{"model_name": k, "quantity": v} for k, v in totals.items()]
    return {"items": items, "total": grand_total}
