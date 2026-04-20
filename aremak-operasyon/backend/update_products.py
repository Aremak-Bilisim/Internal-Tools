"""
1. TIS products: remove * from ProdModel in TeamGram (rows <= 2069)
2. Arducam EK-code matching fix + re-run matching Excel
3. Paraşüt: update name (if corrected) + copy TG Sku as stok kodu for approved matches
"""
import asyncio, re, sys, time
import openpyxl
import httpx
sys.path.insert(0, '.')
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from app.services.teamgram import DOMAIN, BASE, HEADERS
from app.services.parasut import _get_token

PARASUT_BASE = "https://api.parasut.com"
PARASUT_COMPANY = "627949"

# ── helpers ──────────────────────────────────────────────────────────────────

async def tg_get_product(product_id):
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.get(f"{BASE}/{DOMAIN}/Products/Get", headers=HEADERS, params={"id": product_id})
        return r.json()

async def tg_edit_product(product):
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.post(f"{BASE}/{DOMAIN}/Products/Edit", headers=HEADERS, json=product)
        return r.json()

def _norm(s):
    return re.sub(r'[\s\u202f\u00a0]+', ' ', (s or '').strip()).lower()

async def ps_load_all_products(token):
    """Fetch all Paraşüt products, return dict: normalized_name -> {id, name, code}."""
    result = {}
    page = 1
    total_pages = None
    while True:
        try:
            async with httpx.AsyncClient(timeout=30) as c:
                r = await c.get(
                    f"{PARASUT_BASE}/v4/{PARASUT_COMPANY}/products",
                    headers={"Authorization": f"Bearer {token}"},
                    params={"page[size]": 25, "page[number]": page}
                )
                data = r.json()
        except Exception as e:
            print(f"  Sayfa {page} hatasi: {e}, atlaniyor")
            page += 1
            await asyncio.sleep(0.5)
            if total_pages and page > total_pages:
                break
            continue
        items = data.get("data", [])
        for p in items:
            name = p["attributes"].get("name") or ""
            result[_norm(name)] = {"id": p["id"], "name": name, "code": p["attributes"].get("code") or ""}
        meta = data.get("meta", {})
        if total_pages is None:
            total_pages = meta.get("total_pages", 1)
        if page >= total_pages:
            break
        page += 1
        await asyncio.sleep(0.15)
    print(f"  Paraşüt ürünleri yüklendi: {len(result)} kayit ({page}/{total_pages} sayfa)")
    return result

async def ps_search_product(token, name):
    """Fallback: search by exact name via API filter, with 429 retry."""
    for attempt in range(3):
        try:
            async with httpx.AsyncClient(timeout=30) as c:
                r = await c.get(
                    f"{PARASUT_BASE}/v4/{PARASUT_COMPANY}/products",
                    headers={"Authorization": f"Bearer {token}"},
                    params={"filter[name]": name, "page[size]": 5}
                )
            if r.status_code == 429:
                await asyncio.sleep(6)
                continue
            data = r.json()
            for p in data.get("data", []):
                if _norm(p["attributes"].get("name", "")) == _norm(name):
                    attrs = p["attributes"]
                    return {"id": p["id"], "name": attrs.get("name") or "", "code": attrs.get("code") or ""}
            return None
        except Exception:
            await asyncio.sleep(2)
    return None

async def ps_update_product(token, product_id, new_name=None, new_code=None):
    attrs = {}
    if new_name: attrs["name"] = new_name
    if new_code is not None: attrs["code"] = new_code
    if not attrs: return True
    for attempt in range(3):
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.patch(
                f"{PARASUT_BASE}/v4/{PARASUT_COMPANY}/products/{product_id}",
                headers={"Authorization": f"Bearer {token}", "Content-Type": "application/vnd.api+json"},
                json={"data": {"id": str(product_id), "type": "products", "attributes": attrs}}
            )
        if r.status_code == 429:
            await asyncio.sleep(6)
            continue
        return r.status_code < 300
    return False


# ── TASK 1: TIS * cleanup in TeamGram ───────────────────────────────────────

async def clean_tis_stars(tis_star_products):
    print(f"\n=== TIS * temizliği: {len(tis_star_products)} ürün ===")
    ok, fail = 0, 0
    for item in tis_star_products:
        prod = await tg_get_product(item["id"])
        old_model = prod.get("ProdModel", "")
        new_model = old_model.replace("*", "").strip()
        if old_model == new_model:
            continue
        prod["ProdModel"] = new_model
        if prod.get("Category") and prod["Category"].get("Id"):
            prod["CategoryId"] = prod["Category"]["Id"]
        result = await tg_edit_product(prod)
        if result.get("Result") or result.get("Id"):
            print(f"  OK {item['id']}: {repr(old_model)} -> {repr(new_model)}")
            ok += 1
        else:
            print(f"  FAIL {item['id']}: {result}")
            fail += 1
        await asyncio.sleep(0.3)
    print(f"TIS temizlik: {ok} OK, {fail} fail")


# ── TASK 2: Paraşüt updates from approved matches ───────────────────────────

async def update_parasut(approved_rows):
    print(f"\n=== Paraşüt güncelleme: {len(approved_rows)} onaylı satır ===")
    token = await _get_token()
    ps_lookup = await ps_load_all_products(token)  # norm_name -> {id, name, code}
    ok, skip, fail = 0, 0, 0

    for row in approved_rows:
        ps_name = str(row.get("Parasut Urun Adi") or "").strip()
        tg_model = str(row.get("TG Model") or "").strip()
        tg_sku = str(row.get("TG Urun Kodu") or "").strip()

        if not ps_name:
            skip += 1
            continue

        # Look up by PS name, then by TG model (in case already renamed)
        ps_entry = ps_lookup.get(_norm(ps_name)) or ps_lookup.get(_norm(tg_model))
        # Fallback: API search if not in preloaded lookup (pagination gaps)
        if not ps_entry:
            ps_entry = await ps_search_product(token, ps_name)
            if not ps_entry and tg_model != ps_name:
                ps_entry = await ps_search_product(token, tg_model)
            await asyncio.sleep(0.15)
        if not ps_entry:
            print(f"  NOT FOUND: {ps_name[:60]}")
            fail += 1
            continue

        pid = ps_entry["id"]
        current_name = ps_entry["name"]
        current_code = ps_entry["code"]

        # Update name only if TG model differs from current Paraşüt name
        new_name = tg_model if tg_model and _norm(tg_model) != _norm(current_name) else None
        new_code = tg_sku if tg_sku and tg_sku != current_code else None

        if not new_name and not new_code:
            print(f"  SKIP {pid}: {ps_name[:40]} | zaten guncel")
            skip += 1
            continue

        updated = await ps_update_product(token, pid, new_name=new_name, new_code=new_code)
        if updated:
            parts = []
            if new_name: parts.append(f"ad->{new_name[:40]}")
            if new_code: parts.append(f"kod->{new_code}")
            print(f"  OK {pid}: {ps_name[:40]} | {', '.join(parts)}")
            ok += 1
        else:
            print(f"  FAIL {pid}: {ps_name[:40]}")
            fail += 1
        await asyncio.sleep(0.2)

    print(f"Paraşüt: {ok} OK, {skip} skip, {fail} fail")


# ── MAIN ─────────────────────────────────────────────────────────────────────

async def main():
    # Load TeamGram Excel for TIS products
    wb_tg = openpyxl.load_workbook('C:/Users/balam/Downloads/Tum urunler_2026-04-18-10-55.xlsx')
    ws_tg = wb_tg.active
    raw_headers = [str(c.value or '') for c in next(ws_tg.iter_rows(min_row=1, max_row=1))]

    tis_star = []
    for row in ws_tg.iter_rows(min_row=2, max_row=2069, values_only=True):
        d = dict(zip(raw_headers, row))
        marka = next((str(v or '').strip() for k, v in d.items() if 'marka' in k.lower()), '')
        model = next((str(v or '').strip() for k, v in d.items() if 'model' in k.lower()), '')
        if marka == 'TIS' and '*' in model:
            tis_star.append({"id": d.get("Id"), "model": model})

    # Load matching Excel for approved rows up to 2069
    wb_match = openpyxl.load_workbook('C:/Users/balam/Desktop/urun_eslestirme.xlsx')
    ws_match = wb_match.active
    match_headers = [str(c.value or '') for c in ws_match[1]]
    approved = []
    for row in ws_match.iter_rows(min_row=2, max_row=2069, values_only=True):
        d = dict(zip(match_headers, row))
        onay = str(d.get('Onay (Evet/Hayir)') or '').strip().lower()
        if onay == 'evet':
            approved.append(d)

    print(f"TIS * ürünleri (<=2069): {len(tis_star)}")
    print(f"Onaylı eşleşmeler (<=2069): {len(approved)}")

    # Run tasks
    await clean_tis_stars(tis_star)
    await update_parasut(approved)

asyncio.run(main())
